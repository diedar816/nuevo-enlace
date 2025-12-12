# -*- coding: utf-8 -*-
import os
import pandas as pd
import numpy as np
from openpyxl.styles import numbers, Alignment

print("== Inicio del script ==")

# -------------------------------------------------------------------
# Rutas
# -------------------------------------------------------------------
BASE_DIR = r"C:\Proyecto TICs\Nuevo Enlace"
ENTRADA  = os.path.join(BASE_DIR, "Tabla Modelo.XLSX")
SALIDA   = os.path.join(BASE_DIR, "Tabla_ Resultado.xlsx")

print(f"Carpeta base: {BASE_DIR}")
print(f"Archivo entrada: {ENTRADA}")

# -------------------------------------------------------------------
# Festivos Colombia 2025 (ejemplo)
# -------------------------------------------------------------------
HOLIDAYS_2025 = np.array([
    "2025-01-01","2025-01-06","2025-03-24","2025-04-17","2025-04-18","2025-05-01",
    "2025-06-02","2025-06-23","2025-06-30","2025-07-20","2025-08-07","2025-08-18",
    "2025-10-13","2025-11-03","2025-11-17","2025-12-08","2025-12-25"
], dtype="datetime64[D]")
HOLIDAYS_2025 = np.unique(HOLIDAYS_2025)

INCLUDE_END = True  # incluir día final en hábiles

# -------------------------------------------------------------------
# Utilidades
# -------------------------------------------------------------------
def detectar_columna(df, candidatos):
    reales = list(df.columns)
    mapa = {c.lower().strip(): c for c in reales}
    for cand in candidatos:
        k = cand.lower().strip()
        if k in mapa:
            return mapa[k]
    for cand in candidatos:
        k = cand.lower().strip()
        for norm, real in mapa.items():
            if k in norm:
                return real
    return None

def parsear_fecha_solo_date(serie):
    """Convierte a datetime.date (sin hora)."""
    dt = pd.to_datetime(serie, errors="coerce", dayfirst=True)
    return dt.dt.date

LONG_RAD = 14
def normalizar_radicado(x):
    """Texto fijo de 14 dígitos, sin notación científica."""
    if pd.isna(x):
        return None
    s = str(x).strip()
    try:
        if "e" in s.lower():
            s = str(int(float(s)))
        else:
            s_num = s.replace(",", "").replace(" ", "")
            if s_num.replace(".", "", 1).isdigit():
                s = str(int(float(s_num)))
    except:
        pass
    s_digits = "".join(ch for ch in s if ch.isdigit())
    return s_digits.zfill(LONG_RAD) if s_digits else s

def normalizar_texto(x):
    """Uppercase sin tildes para dependencias."""
    if pd.isna(x):
        return ""
    s = str(x).upper()
    return (s.replace("Á","A").replace("É","E")
             .replace("Í","I").replace("Ó","O")
             .replace("Ú","U"))

def dias_habiles(a, b, include_end=INCLUDE_END):
    """Días hábiles (lun-vie) entre a y b excluyendo festivos (2025)."""
    if a is None or b is None or pd.isna(a) or pd.isna(b):
        return np.nan
    start = np.datetime64(a)  # a y b son datetime.date
    end   = np.datetime64(b)

    if include_end and end >= start:
        end += np.timedelta64(1, "D")
    elif include_end and end < start:
        start += np.timedelta64(1, "D")
        return -np.busday_count(end, start, weekmask="1111100", holidays=HOLIDAYS_2025)

    if end >= start:
        return np.busday_count(start, end, weekmask="1111100", holidays=HOLIDAYS_2025)
    else:
        return -np.busday_count(end, start, weekmask="1111100", holidays=HOLIDAYS_2025)

def dias_naturales(a, b, inclusive=False):
    """Días naturales entre a y b. Si inclusive=True, suma 1."""
    if a is None or b is None or pd.isna(a) or pd.isna(b):
        return np.nan
    diff = (b - a).days
    return diff + 1 if inclusive else diff

# --- Normalización de nombres y deduplicación conservando la última (derecha) ---
def norm_colname(s: str) -> str:
    return str(s).strip().upper().replace(" ", "_")

def dedupe_keep_last_by_name(df: pd.DataFrame, target: str) -> pd.DataFrame:
    """Si hay varias columnas cuyo nombre normalizado coincide con target, conserva la última y elimina anteriores."""
    target_norm = norm_colname(target)
    cols = list(df.columns)
    idxs = [i for i, c in enumerate(cols) if norm_colname(c) == target_norm]
    if len(idxs) <= 1:
        return df
    drop_idx_set = set(idxs[:-1])
    new_cols = [c for i, c in enumerate(cols) if i not in drop_idx_set]
    return df[new_cols]

# -------------------------------------------------------------------
# Lectura
# -------------------------------------------------------------------
if not os.path.exists(ENTRADA):
    raise FileNotFoundError(f"No se encontró el archivo: {ENTRADA}")

print("Leyendo Excel...")
df = pd.read_excel(ENTRADA, engine="openpyxl")
print(f"Filas: {len(df)} | Columnas: {len(df.columns)}")

# -------------------------------------------------------------------
# Detección de columnas principales
# -------------------------------------------------------------------
col_radicado  = detectar_columna(df, ["numero_radicado","nro_radicado","radicado"])
col_dep       = detectar_columna(df, ["nombre_dependencia","dependencia"])
col_asigna    = detectar_columna(df, ["fecha_asignacion","fecha_asigna","f_asignacion","asignacion"])
col_caso_crea = detectar_columna(df, ["CASO_FECHA_CREACION","caso_fecha_creacion","fecha_creacion_caso","f_caso_crea"])

print("Columnas detectadas:")
print("  numero_radicado     =>", col_radicado)
print("  nombre_dependencia  =>", col_dep)
print("  fecha_asignacion    =>", col_asigna)
print("  CASO_FECHA_CREACION =>", col_caso_crea)

faltantes = [k for k,v in {
    "numero_radicado": col_radicado,
    "nombre_dependencia": col_dep,
    "fecha_asignacion": col_asigna,
    "CASO_FECHA_CREACION": col_caso_crea,
}.items() if v is None]
if faltantes:
    raise ValueError(f"Faltan columnas imprescindibles: {faltantes}")

# -------------------------------------------------------------------
# Normalizar fechas (sin hora) y radicado (texto de 14 dígitos)
# -------------------------------------------------------------------
df[col_asigna]    = parsear_fecha_solo_date(df[col_asigna])
df[col_caso_crea] = parsear_fecha_solo_date(df[col_caso_crea])

df["_fecha_asignacion"]    = df[col_asigna]          # internas (date)
df["_caso_fecha_creacion"] = df[col_caso_crea]       # internas (date)
df["_rad_str"]             = df[col_radicado].apply(normalizar_radicado).astype("string")

# Usar radicado normalizado hacia adelante
df[col_radicado] = df["_rad_str"].astype("string")

# -------------------------------------------------------------------
# Máscaras de dependencia robustas
# -------------------------------------------------------------------
dep_norm = df[col_dep].apply(normalizar_texto)
mask_vent = dep_norm.str.contains("VENTANILL", na=False)   # captura VENTANILLA/VENTANILL
mask_mesa = dep_norm.str.contains("MESA", na=False)

# -------------------------------------------------------------------
# PAREO VERTICAL (VENTANILLA ↔ MESA) POR RADICADO
# -------------------------------------------------------------------
print("Construyendo referencia de fecha_asignacion en MESA por radicado...")
mesa_asig = (df.loc[mask_mesa, [col_radicado, "_fecha_asignacion"]]
               .dropna()
               .groupby(col_radicado)["_fecha_asignacion"]
               .apply(list)
               .to_dict())

def mesa_ref_date(radicado, vent_date):
    """Primera fecha MESA >= vent_date; si no hay, mínima MESA."""
    fechas = mesa_asig.get(radicado, [])
    fechas_validas = [d for d in fechas if pd.notna(d)]
    if not fechas_validas:
        return pd.NaT
    posteriores = [d for d in fechas_validas if d >= vent_date]
    return min(posteriores) if posteriores else min(fechas_validas)

print("Calculando TOTAL_DIAS_VENTANILLA (hábiles y naturales)...")
df["TOTAL_DIAS_VENTANILLA_NATURALES"] = np.nan  # interno
df["TOTAL_DIAS_VENTANILLA_HABILES"]   = np.nan  # interno

pareos_logrados = 0

for i in df.index:
    if mask_vent.iloc[i] and pd.notna(df.loc[i, "_fecha_asignacion"]):
        rad = df.loc[i, col_radicado]
        vent_date = df.loc[i, "_fecha_asignacion"]
        mesa_date = mesa_ref_date(rad, vent_date)
        if pd.notna(mesa_date):
            pareos_logrados += 1
            df.at[i, "TOTAL_DIAS_VENTANILLA_NATURALES"] = dias_naturales(vent_date, mesa_date, inclusive=False)
            df.at[i, "TOTAL_DIAS_VENTANILLA_HABILES"]   = dias_habiles(vent_date, mesa_date, include_end=INCLUDE_END)

# Principal = hábiles
df["TOTAL_DIAS_VENTANILLA"] = df["TOTAL_DIAS_VENTANILLA_HABILES"]
print(f"Pareos VENT ↔ MESA logrados: {pareos_logrados}/{int(mask_vent.sum())}")

# -------------------------------------------------------------------
# HORIZONTAL EN MESA
# -------------------------------------------------------------------
print("Calculando TIEMPO_EN_MESA_DE_CREACION (hábiles y naturales)...")
df["TIEMPO_EN_MESA_DE_CREACION_NATURALES"] = np.nan  # interno (oculto)
df["TIEMPO_EN_MESA_DE_CREACION_HABILES"]   = np.nan  # interno (oculto)

mesa_rows = df.loc[mask_mesa].index
df.loc[mesa_rows, "TIEMPO_EN_MESA_DE_CREACION_NATURALES"] = [
    dias_naturales(a, b, inclusive=False) for a, b in zip(
        df.loc[mesa_rows, "_fecha_asignacion"],
        df.loc[mesa_rows, "_caso_fecha_creacion"]
    )
]
df.loc[mesa_rows, "TIEMPO_EN_MESA_DE_CREACION_HABILES"] = [
    dias_habiles(a, b, include_end=INCLUDE_END) for a, b in zip(
        df.loc[mesa_rows, "_fecha_asignacion"],
        df.loc[mesa_rows, "_caso_fecha_creacion"]
    )
]
df["TIEMPO_EN_MESA_DE_CREACION"] = df["TIEMPO_EN_MESA_DE_CREACION_HABILES"]  # principal

# -------------------------------------------------------------------
# ALERTAS (visible)
# -------------------------------------------------------------------
print("Generando alertas...")
ALERTA = []
for i, row in df.iterrows():
    avisos = []
    if mask_vent.iloc[i]:
        if pd.isna(row["TOTAL_DIAS_VENTANILLA_HABILES"]) and pd.notna(row["_fecha_asignacion"]):
            avisos.append("VENT: Sin MESA para pareo (mismo radicado)")
        if pd.notna(row["TOTAL_DIAS_VENTANILLA_HABILES"]) and row["TOTAL_DIAS_VENTANILLA_HABILES"] < 0:
            avisos.append("VENT: Diferencia negativa (MESA < VENT)")
        cant_mesa = len([d for d in mesa_asig.get(row[col_radicado], []) if pd.notna(d)])
        if cant_mesa > 1 and pd.notna(row["TOTAL_DIAS_VENTANILLA_HABILES"]):
            avisos.append(f"VENT: MESA múltiple ({cant_mesa}) → se eligió la más cercana ≥ VENT")
    if mask_mesa.iloc[i]:
        if pd.isna(row["_fecha_asignacion"]) or pd.isna(row["_caso_fecha_creacion"]):
            avisos.append("MESA: Fecha(s) faltante(s)")
        elif pd.notna(row["TIEMPO_EN_MESA_DE_CREACION_HABILES"]) and row["TIEMPO_EN_MESA_DE_CREACION_HABILES"] < 0:
            avisos.append("MESA: Diferencia negativa (CREACIÓN < ASIGNACIÓN)")
    ALERTA.append("; ".join(avisos) if avisos else "")
df["ALERTA"] = ALERTA

# -------------------------------------------------------------------
# Métricas de calidad (explican por qué salen alertas de fechas faltantes)
# -------------------------------------------------------------------
faltan_asig_mesa = int(df.loc[mask_mesa, "_fecha_asignacion"].isna().sum())
faltan_crea_mesa = int(df.loc[mask_mesa, "_caso_fecha_creacion"].isna().sum())
faltan_asig_vent = int(df.loc[mask_vent, "_fecha_asignacion"].isna().sum())
print(f"Fechas faltantes en MESA → asignación: {faltan_asig_mesa} | creación: {faltan_crea_mesa}")
print(f"Fechas faltantes en VENTANILLA → asignación: {faltan_asig_vent}")

# -------------------------------------------------------------------
# Filas pertinentes
# -------------------------------------------------------------------
pertinentes = df["ALERTA"].astype(bool) | \
              df["TOTAL_DIAS_VENTANILLA"].notna() | \
              df["TIEMPO_EN_MESA_DE_CREACION"].notna()
df_pert = df.loc[pertinentes].copy()

# -------------------------------------------------------------------
# Resumen por radicado (vista rápida) - SOLO columnas principales
# -------------------------------------------------------------------
print("Construyendo resumen por radicado...")
tmp = df.loc[mask_vent | mask_mesa, [col_radicado, col_dep,
    "TOTAL_DIAS_VENTANILLA","TIEMPO_EN_MESA_DE_CREACION"]].copy()
tmp["tipo"] = np.where(tmp[col_dep].astype(str).str.upper().str.startswith("VENTANILL"), "VENTANILLA", "MESA")

summary = (tmp
    .pivot_table(index=col_radicado,
                 columns="tipo",
                 values=["TOTAL_DIAS_VENTANILLA","TIEMPO_EN_MESA_DE_CREACION"],
                 aggfunc="max")
    .reset_index()
)

# Asegurar que la 1ª columna sea exactamente col_radicado
first_col = summary.columns[0]
if first_col != col_radicado:
    summary.rename(columns={first_col: col_radicado}, inplace=True)

summary.columns = [
    (c if isinstance(c, str) else f"{c[0]}_{c[1]}")
    for c in summary.columns
]

# -------------------------------------------------------------------
# Auditoría columnas detectadas
# -------------------------------------------------------------------
detalles_cols = pd.DataFrame([
    {"campo_logico": "numero_radicado", "columna_en_archivo": col_radicado},
    {"campo_logico": "nombre_dependencia", "columna_en_archivo": col_dep},
    {"campo_logico": "fecha_asignacion", "columna_en_archivo": col_asigna},
    {"campo_logico": "CASO_FECHA_CREACION", "columna_en_archivo": col_caso_crea},
    {"campo_logico": "Regla TOTAL_DIAS_VENTANILLA", "columna_en_archivo": "MESA.fecha_asignacion - VENT.fecha_asignacion (mismo radicado)"},
    {"campo_logico": "Regla TIEMPO_EN_MESA_DE_CREACION", "columna_en_archivo": "MESA.CASO_FECHA_CREACION - MESA.fecha_asignacion"}
])

# -------------------------------------------------------------------
# Auditoría de pareo (nueva hoja)
# -------------------------------------------------------------------
print("Construyendo hoja Auditoria_pareo...")
aud_rows = []
vent_rows = df.loc[mask_vent].index
for i in vent_rows:
    rad = df.loc[i, col_radicado]
    vent_date = df.loc[i, "_fecha_asignacion"]
    mesas = mesa_asig.get(rad, [])
    mesas_validas = [d for d in mesas if pd.notna(d)]
    mesa_sel = mesa_ref_date(rad, vent_date) if pd.notna(vent_date) else pd.NaT

    if not mesas_validas:
        estado = "SIN_PAREO"
        motivo = "Sin fechas MESA"
    elif pd.isna(vent_date):
        estado = "SIN_PAREO"
        motivo = "VENT sin fecha"
    elif pd.notna(mesa_sel):
        if mesa_sel >= vent_date:
            estado = "PAREO"
            motivo = "MESA ≥ VENT"
        else:
            estado = "PAREO"
            motivo = "MESA mínima < VENT"
    else:
        estado = "SIN_PAREO"
        motivo = "No se encontró MESA ≥ VENT"

    aud_rows.append({
        col_radicado: rad,
        "VENT_FECHA_ASIGNACION": vent_date,
        "MESA_FECHAS_DISPONIBLES": ", ".join(sorted([d.strftime("%Y-%m-%d") for d in mesas_validas])) if mesas_validas else "",
        "MESA_FECHA_SELECCIONADA": mesa_sel,
        "ESTADO_PAREO": estado,
        "MOTIVO": motivo
    })

auditoria_pareo = pd.DataFrame(aud_rows)

# -------------------------------------------------------------------
# Construir dataframes de salida: ocultar técnicas y ordenar visibles
# -------------------------------------------------------------------
drop_cols = [
    # Internas por prefijo "_"
    *[c for c in df.columns if c.startswith("_")],
    # Desgloses técnicos (no mostrar)
    "TOTAL_DIAS_VENTANILLA_HABILES",
    "TOTAL_DIAS_VENTANILLA_NATURALES",
    "TIEMPO_EN_MESA_DE_CREACION_HABILES",
    "TIEMPO_EN_MESA_DE_CREACION_NATURALES",
    # Si existe una columna literal "N" (auxiliar), eliminarla
    "N",
]
df_out = df.drop(columns=drop_cols, errors="ignore")
df_pert_out = df_pert.drop(columns=drop_cols, errors="ignore")

# --- Deduplicar: conservar SOLO la última 'TOTAL_DIAS_VENTANILLA' (la de la derecha) ---
df_out = dedupe_keep_last_by_name(df_out,  "TOTAL_DIAS_VENTANILLA")
df_pert_out = dedupe_keep_last_by_name(df_pert_out, "TOTAL_DIAS_VENTANILLA")

# Reordenar: dejar al final las columnas principales (ALERTA, TIEMPO_EN_MESA_DE_CREACION, TOTAL_DIAS_VENTANILLA)
final_cols = ["ALERTA", "TIEMPO_EN_MESA_DE_CREACION", "TOTAL_DIAS_VENTANILLA"]
for c in final_cols:
    if c in df_out.columns:
        df_out = df_out[[col for col in df_out.columns if col != c] + [c]]
    if c in df_pert_out.columns:
        df_pert_out = df_pert_out[[col for col in df_pert_out.columns if col != c] + [c]]

# -------------------------------------------------------------------
# Guardar Excel con formato de fecha, texto y alineación solicitada
# -------------------------------------------------------------------
print("Guardando Excel de salida...")
with pd.ExcelWriter(SALIDA, engine="openpyxl", datetime_format="YYYY-MM-DD", date_format="YYYY-MM-DD") as wr:
    df_out.to_excel(wr, index=False, sheet_name="Todos")
    df_pert_out.to_excel(wr, index=False, sheet_name="Registros_para_revisar")
    summary.to_excel(wr, index=False, sheet_name="Resumen_rad_duplicados")
    detalles_cols.to_excel(wr, index=False, sheet_name="Columnas_detectadas")
    auditoria_pareo.to_excel(wr, index=False, sheet_name="Auditoria_pareo")

    wb = wr.book

    # === Formateo de fechas (sin hora) ===
    fmt_fecha = numbers.FORMAT_DATE_YYYYMMDD2  # yyyy-mm-dd

    def formatear_fechas(ws, nombres_columnas):
        header_idx = {cell.value: cell.column for cell in ws[1] if cell.value in nombres_columnas}
        for col_name, col_index in header_idx.items():
            col_letter = ws.cell(row=1, column=col_index).column_letter
            for r in range(2, ws.max_row + 1):
                ws[f"{col_letter}{r}"].number_format = fmt_fecha

    # === Alineación solicitada (todas las celdas) ===
    # Horizontal: izquierda; Vertical: superior; Justificado/Distribuido y wrap_text.
    align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_distributed_top = Alignment(horizontal="distributed", vertical="top", wrap_text=True)

    def alinear_hoja(ws):
        # Aplica alineación a todas las celdas con dos pasadas:
        # 1) izquierda-arriba (base)
        # 2) distribuido-arriba (mejor justificado visual)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align_left_top
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align_distributed_top

    # Hoja "Todos": formatear fechas y alinear
    ws = wb["Todos"]
    headers_todos = [cell.value for cell in ws[1]]
    fecha_cols_todos = [col_asigna, col_caso_crea]
    formatear_fechas(ws, set(fecha_cols_todos) & set(headers_todos))
    alinear_hoja(ws)

    # Hoja "Registros_para_revisar": fechas y alineación
    ws = wb["Registros_para_revisar"]
    headers_rev = [cell.value for cell in ws[1]]
    fecha_cols_rev = [col_asigna, col_caso_crea]
    formatear_fechas(ws, set(fecha_cols_rev) & set(headers_rev))
    alinear_hoja(ws)

    # Hoja "Resumen_rad_duplicados": alineación general
    ws = wb["Resumen_rad_duplicados"]
    alinear_hoja(ws)

    # Hoja "Columnas_detectadas": alineación general
    ws = wb["Columnas_detectadas"]
    alinear_hoja(ws)

    # Hoja "Auditoria_pareo": fechas y alineación
    ws = wb["Auditoria_pareo"]
    headers_aud = [cell.value for cell in ws[1]]
    fecha_cols_aud = ["VENT_FECHA_ASIGNACION", "MESA_FECHA_SELECCIONADA"]
    formatear_fechas(ws, set(fecha_cols_aud) & set(headers_aud))
    alinear_hoja(ws)

    # === Forzar numero_radicado como texto en Excel (prefijo apóstrofe) ===
    for sheet_name in ["Todos", "Resumen_rad_duplicados", "Auditoria_pareo"]:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if col_radicado in headers:
            # buscar letra de columna
            for cell in ws[1]:
                if cell.value == col_radicado:
                    col_letter = cell.column_letter
                    break
            for r in range(2, ws.max_row + 1):
                val = ws[f"{col_letter}{r}"].value
                if val is not None:
                    ws[f"{col_letter}{r}"].value = "'" + str(val)

